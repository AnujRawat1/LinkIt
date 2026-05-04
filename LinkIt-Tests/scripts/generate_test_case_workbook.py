from pathlib import Path
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
REPORT_FILE = ROOT / "target" / "surefire-reports" / "testng-results.xml"
OUTPUT_DIR = ROOT / "docs"
OUTPUT_FILE = OUTPUT_DIR / "LinkIt-Test-Cases.xlsx"

COLUMNS = [
    "Test Case ID",
    "Summary",
    "Description",
    "Pre-conditions",
    "Test Steps",
    "Expected Result",
    "Status(Pass/Failed)",
]

API_CASES = [
    {
        "method": "createRoomShouldReturnDefaults",
        "id": "API-001",
        "summary": "Create room returns default room details",
        "description": "Verify that creating a room returns a 6-character room ID, default editor content, initialized file list, and the creator in participants.",
        "pre": "Backend is running on localhost:8080. MongoDB Atlas connection is valid. API endpoint /api/room/createRoom is available.",
        "steps": "1. Send POST request to /api/room/createRoom with a valid name.\n2. Capture the response body.\n3. Validate roomId, content, participants, and fileNames.",
        "expected": "The API returns HTTP 200 with a non-empty 6-character roomId, default editor text, an initialized fileNames list, and the creator included in participants.",
    },
    {
        "method": "roomIdsShouldBeUniqueAcrossCreates",
        "id": "API-002",
        "summary": "Two room creation requests generate unique IDs",
        "description": "Verify that consecutive room creation requests do not reuse the same room ID.",
        "pre": "Backend is running and room creation API is accessible.",
        "steps": "1. Send POST request to create the first room.\n2. Send POST request to create the second room.\n3. Compare both returned room IDs.",
        "expected": "Both requests return HTTP 200 and the generated room IDs are different.",
    },
    {
        "method": "joinRoomShouldAddParticipant",
        "id": "API-003",
        "summary": "Join room adds a new participant",
        "description": "Verify that a valid join request adds the guest user to the room participants list.",
        "pre": "A room already exists with one owner participant.",
        "steps": "1. Create a room with an owner name.\n2. Send POST request to /api/room/joinRoom with the same roomId and a guest name.\n3. Validate the participants list in the response.",
        "expected": "The API returns HTTP 200 and the response participants list contains both the owner and the guest.",
    },
    {
        "method": "joinInvalidRoomShouldReturnNoContent",
        "id": "API-004",
        "summary": "Join invalid room returns no content",
        "description": "Verify that joining a non-existent room is rejected without room data.",
        "pre": "Backend is running. A fake room ID that does not exist is available.",
        "steps": "1. Send POST request to /api/room/joinRoom with an invalid roomId and valid name.\n2. Capture the status code.",
        "expected": "The API returns HTTP 204 and no room object is returned.",
    },
    {
        "method": "getContentShouldReturnDefaultContent",
        "id": "API-005",
        "summary": "Get content returns default text for a new room",
        "description": "Verify that fetching content from a newly created room returns the default editor template.",
        "pre": "A room has been created successfully.",
        "steps": "1. Create a new room.\n2. Send GET request to /api/room/getContent with the created roomId.\n3. Validate the response body.",
        "expected": "The API returns HTTP 200 and the response body matches the default editor content string.",
    },
    {
        "method": "getParticipantsShouldReflectJoins",
        "id": "API-006",
        "summary": "Get participants reflects room joins",
        "description": "Verify that the participants endpoint returns all users currently in the room after a join operation.",
        "pre": "A room exists and at least one additional participant has joined it.",
        "steps": "1. Create a room with an owner.\n2. Join the same room using a guest user.\n3. Send GET request to /api/room/getParticipants with the roomId.\n4. Validate the participant names returned.",
        "expected": "The API returns HTTP 200 and the participants list contains both owner and guest names.",
    },
    {
        "method": "getFileNamesShouldStartEmpty",
        "id": "API-007",
        "summary": "Get file names starts with empty list",
        "description": "Verify that a newly created room has no uploaded file names associated with it.",
        "pre": "A room exists and no files have been added to it.",
        "steps": "1. Create a new room.\n2. Send GET request to /api/room/getFileNames with the roomId.\n3. Inspect the returned list.",
        "expected": "The API returns HTTP 200 and the file names list is empty.",
    },
    {
        "method": "removeLastParticipantShouldDeleteRoom",
        "id": "API-008",
        "summary": "Removing last participant deletes room",
        "description": "Verify that a room is deleted when its final participant leaves.",
        "pre": "A room exists with exactly one participant.",
        "steps": "1. Create a room with a single participant.\n2. Send DELETE request to /api/room/removeParticipant for that participant.\n3. Send GET request to /api/room/getContent for the same roomId.\n4. Compare the returned status codes.",
        "expected": "The remove request returns HTTP 200, and the subsequent content request returns HTTP 204 because the room no longer exists.",
    },
    {
        "method": "removeParticipantFromUnknownRoomShouldReturnNoContent",
        "id": "API-009",
        "summary": "Remove participant from unknown room returns no content",
        "description": "Verify that removeParticipant gracefully handles an invalid room ID.",
        "pre": "Backend is running. A room ID that does not exist is available.",
        "steps": "1. Send DELETE request to /api/room/removeParticipant with an unknown roomId and any participantName.\n2. Capture the status code.",
        "expected": "The API returns HTTP 204 when the specified room does not exist.",
    },
    {
        "method": "getParticipantsForUnknownRoomShouldReturnNoContent",
        "id": "API-010",
        "summary": "Get participants for unknown room returns no content",
        "description": "Verify that the participants endpoint does not return data for a non-existent room.",
        "pre": "Backend is running. A room ID that does not exist is available.",
        "steps": "1. Send GET request to /api/room/getParticipants with an unknown roomId.\n2. Capture the status code.",
        "expected": "The API returns HTTP 204 and no participants list is returned.",
    },
]

UI_CASES = [
    {
        "method": "homePageShouldRenderPrimaryControls",
        "id": "UI-001",
        "summary": "Homepage shows main create and join controls",
        "description": "Verify that the landing page renders the Create tab, Join tab, and name input field.",
        "pre": "Frontend is running on localhost:5173 and accessible in a browser.",
        "steps": "1. Open the LinkIt homepage.\n2. Observe the available tabs and form fields.\n3. Confirm presence of the name input box.",
        "expected": "The homepage displays the Create tab, Join tab, and name input field successfully.",
    },
    {
        "method": "joinTabShouldRevealRoomIdField",
        "id": "UI-002",
        "summary": "Join tab reveals room ID field",
        "description": "Verify that switching to Join mode displays the room ID input field.",
        "pre": "Frontend is running and homepage is open.",
        "steps": "1. Open the LinkIt homepage.\n2. Click the Join Clipboard tab.\n3. Check whether the room ID input field appears.",
        "expected": "The room ID input field becomes visible after the Join tab is selected.",
    },
    {
        "method": "createRoomFlowShouldOpenRoomPage",
        "id": "UI-003",
        "summary": "Create room flow navigates to room page",
        "description": "Verify that creating a room from the homepage redirects the user to the collaborative room screen.",
        "pre": "Frontend and backend are running. Room creation API is reachable.",
        "steps": "1. Open the homepage.\n2. Enter a valid user name.\n3. Click Create Clipboard.\n4. Observe the redirected page URL and room details.",
        "expected": "The application navigates to /room/{roomId}, shows a room ID badge, and lists the creator in participants.",
    },
    {
        "method": "joinExistingRoomShouldNavigateToRoom",
        "id": "UI-004",
        "summary": "Join existing room navigates to selected room",
        "description": "Verify that a user can join an existing room from the homepage and land on the correct room route.",
        "pre": "A valid room already exists. Frontend and backend are running.",
        "steps": "1. Create a room through the API or UI.\n2. Open the homepage.\n3. Switch to Join tab.\n4. Enter a guest name and the valid room ID.\n5. Click Join Clipboard.\n6. Observe the redirected URL and participants list.",
        "expected": "The application navigates to the selected room URL and the participants list includes the joined guest.",
    },
    {
        "method": "invalidJoinShouldShowErrorToast",
        "id": "UI-005",
        "summary": "Invalid join shows error toast",
        "description": "Verify that attempting to join an invalid room keeps the user on the homepage and shows an error message.",
        "pre": "Frontend and backend are running. An invalid room ID is available.",
        "steps": "1. Open the homepage.\n2. Switch to Join tab.\n3. Enter a valid user name and an invalid room ID.\n4. Click Join Clipboard.\n5. Observe the toast notification and current URL.",
        "expected": "An error toast stating that the room does not exist is shown, and the user remains on the homepage instead of navigating to a room.",
    },
    {
        "method": "deepLinkRoomRouteShouldLoadWithUrlParam",
        "id": "UI-006",
        "summary": "Deep link room route loads using URL parameter",
        "description": "Verify that opening the room page directly with /room/{roomId} displays the correct room ID on screen.",
        "pre": "A valid room already exists. Frontend and backend are running.",
        "steps": "1. Create a room and note its room ID.\n2. Open the browser directly on /room/{roomId}.\n3. Observe the room badge shown in the UI.",
        "expected": "The room screen loads successfully and the displayed room ID matches the value in the URL.",
    },
]


def parse_statuses(report_file: Path):
    statuses = {}
    executed_at = ""
    if not report_file.exists():
        return statuses, executed_at

    root = ET.parse(report_file).getroot()
    suite = root.find("suite")
    if suite is not None:
        executed_at = suite.attrib.get("finished-at", "")

    for class_node in root.findall(".//class"):
        class_name = class_node.attrib.get("name", "")
        for method_node in class_node.findall("test-method"):
            if method_node.attrib.get("is-config") == "true":
                continue
            method_name = method_node.attrib.get("name", "")
            status = method_node.attrib.get("status", "NOT RUN")
            statuses[(class_name, method_name)] = status
    return statuses, executed_at


def normalize_status(raw_status: str):
    mapping = {
        "PASS": "Pass",
        "FAILED": "Failed",
        "FAIL": "Failed",
        "SKIP": "Not Run",
        "SKIPPED": "Not Run",
        "NOT RUN": "Not Run",
    }
    return mapping.get(raw_status.upper(), raw_status.title() if raw_status else "Not Run")


def apply_sheet_style(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    wrap = Alignment(vertical="top", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    widths = {
        1: 16,
        2: 36,
        3: 52,
        4: 45,
        5: 65,
        6: 55,
        7: 18,
    }
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"


def add_case_rows(ws, cases, class_name, statuses):
    ws.append(COLUMNS)
    for case in cases:
        status = normalize_status(statuses.get((class_name, case["method"]), "Not Run"))
        ws.append([
            case["id"],
            case["summary"],
            case["description"],
            case["pre"],
            case["steps"],
            case["expected"],
            status,
        ])
    apply_sheet_style(ws)


def build_workbook():
    statuses, executed_at = parse_statuses(REPORT_FILE)
    documented_statuses = []
    for case in API_CASES:
        documented_statuses.append(statuses.get(("com.linkit.tests.api.RoomApiTests", case["method"]), "NOT RUN"))
    for case in UI_CASES:
        documented_statuses.append(statuses.get(("com.linkit.tests.ui.RoomUiTests", case["method"]), "NOT RUN"))

    wb = Workbook()

    overview = wb.active
    overview.title = "Overview"
    overview.append(["Item", "Value"])
    overview.append(["Project", "LinkIt"])
    overview.append(["Workbook Purpose", "Document backend API and frontend UI automated test cases in Excel format."])
    overview.append(["Execution Source", str(REPORT_FILE.relative_to(ROOT)) if REPORT_FILE.exists() else "No execution report found"])
    overview.append(["Last Verified", executed_at or "Not Run"])
    overview.append(["API Test Cases", len(API_CASES)])
    overview.append(["UI Test Cases", len(UI_CASES)])
    overview.append(["Executed Business Test Cases", sum(1 for status in documented_statuses if status in {"PASS", "FAIL", "FAILED", "SKIP", "SKIPPED"})])
    overview.append(["Passed Business Test Cases", sum(1 for status in documented_statuses if status == "PASS")])
    overview.append(["Failed Business Test Cases", sum(1 for status in documented_statuses if status in {"FAIL", "FAILED"})])
    overview.append(["Status Note", "Status values are derived from actual TestNG/Surefire execution results."])
    apply_sheet_style(overview)
    overview.column_dimensions["A"].width = 28
    overview.column_dimensions["B"].width = 90

    api_sheet = wb.create_sheet("API Test Cases")
    add_case_rows(api_sheet, API_CASES, "com.linkit.tests.api.RoomApiTests", statuses)

    ui_sheet = wb.create_sheet("UI Test Cases")
    add_case_rows(ui_sheet, UI_CASES, "com.linkit.tests.ui.RoomUiTests", statuses)

    for sheet in (api_sheet, ui_sheet):
        for row in range(2, sheet.max_row + 1):
            status_cell = sheet.cell(row=row, column=7)
            if status_cell.value == "Pass":
                status_cell.fill = PatternFill("solid", fgColor="C6EFCE")
                status_cell.font = Font(color="006100", bold=True)
            elif status_cell.value == "Failed":
                status_cell.fill = PatternFill("solid", fgColor="FFC7CE")
                status_cell.font = Font(color="9C0006", bold=True)
            else:
                status_cell.fill = PatternFill("solid", fgColor="FFEB9C")
                status_cell.font = Font(color="9C6500", bold=True)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE


if __name__ == "__main__":
    output = build_workbook()
    print(output)

